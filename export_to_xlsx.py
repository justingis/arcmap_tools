#-------------------------------------------------------------------------------
# Name:        Export to XLSX
# Purpose:     Exports a layer or table to .XLSX format. User may select which fields to include in the export.
#
# Author:      Justin Hawley (justin@orcagis.com)
#
# Created:     06/13/2022
#
# Interpreter: C:\Python27\ArcGIS10.8\python.exe
#-------------------------------------------------------------------------------

import arcpy
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter


class clsField(object):
    """ Class to hold properties and behavior of the output fields
    """
    @property
    def alias(self):
        return self._field.aliasName

    @property
    def name(self):
        return self._field.name

    @property
    def domain(self):
        return self._field.domain

    @property
    def type(self):
        return self._field.type

    @property
    def length(self):
        return self._field.length

    def __init__(self, f, i, subtypes):
        """ Create the object from a describe field object
        """
        self.index = None
        self._field = f
        self.subtype_field = ''
        self.domain_desc = {}
        self.subtype_desc = {}
        self.index = i

        for st_key, st_val in subtypes.iteritems():
            if st_val['SubtypeField'] == f.name:
                self.subtype_desc[st_key] = st_val['Name']
                self.subtype_field = f.name
            for k, v in st_val['FieldValues'].iteritems():
                if k == f.name:
                    if len(v) == 2:
                        if v[1]:
                            self.domain_desc[st_key]= v[1].codedValues
                            self.subtype_field = st_val['SubtypeField']

    def __repr__(self):
        """ Nice representation for debugging  """
        return '<clsfield object name={}, alias={}, domain_desc={}>'.format(self.name,
                                                                self.alias,
                                                                self.domain_desc)

    def updateValue(self, row, fields):
        """ Update value based on domain/subtypes """
        value = row[self.index]
        if self.subtype_field:
            subtype_val = row[fields.index(self.subtype_field)]
        else:
            subtype_val = 0

        if self.subtype_desc:
            value = self.subtype_desc[row[self.index]]

        if self.domain_desc:
            try:
                value = self.domain_desc[subtype_val][row[self.index]]
            except:
                pass # not all subtypes will have domain

        return value

def get_field_defs(in_table, use_domain_desc):
    desc = arcpy.Describe(in_table)

    subtypes ={}
    if use_domain_desc:
        subtypes = arcpy.da.ListSubtypes(in_table)

    fields = []
    for i, field in enumerate([f for f in desc.fields
                                if f.type in ["Date","Double","Guid",
                                              "Integer","OID","Single",
                                              "SmallInteger","String"]]):
        fields.append(clsField(field, i, subtypes))

    return fields



       
def table_to_excel(in_table, output, use_field_alias=False, use_domain_desc=False):

    fieldNames_forExcel = []
    wb = Workbook()
    ws = wb.active
    #font = Font(name='Calibri',size=11,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
    
    arcpy.AddMessage("Creating Excel file here: " + output)

    fields = get_field_defs(in_table, use_domain_desc)
    field_names = [i.name for i in fields]

    
    checkedFields = arcpy.GetParameter(4)
    stringCheckedFields = []


    arcpy.AddMessage("\n")

    if len(checkedFields) > 0:
        for check in checkedFields:
            stringCheckedFields.append(str(check))
        for field in fields:
            if (use_field_alias == "true"):
                if str(field.name) in stringCheckedFields:
                    fieldNames_forExcel.append(field.alias)
            else:
                if str(field.name) in stringCheckedFields:
                    fieldNames_forExcel.append(field.name)
    else:
        stringCheckedFields = field_names
        if use_field_alias == True:
            fieldNames_forExcel = [i.alias for i in fields]
        else:
            fieldNames_forExcel = [i.name for i in fields]


    inputDesc = arcpy.Describe(in_table)
    sheetName = arcpy.GetParameterAsText(5)
    if sheetName == "":
        ws.title = inputDesc.name
    else:
        ws.title = sheetName
    
    ws.append(fieldNames_forExcel)
    
    columnList = list(ws.column_dimensions)
    
    excelFieldList = list(ws)
    
    for excelField in excelFieldList:
        for field in excelField:
            field.font = Font(bold=True)
            field.alignment=Alignment(horizontal='center')
            column_letter = get_column_letter(field.column)
            ws.column_dimensions[column_letter].width = 25 #may need some work
            
    
    with arcpy.da.SearchCursor(in_table, stringCheckedFields) as cursor: #(field_names)
        for row in cursor:
            dataRowList = []
            for col_index, value in enumerate(row):
                if (fields[col_index].domain_desc or fields[col_index].subtype_desc):
                    #value = fields[col_index].updateValue(row, field_names)
                    value = fields[col_index].updateValue(row, stringCheckedFields)
                dataRowList.append(value)
            ws.append(dataRowList)
    
    
    #wb.save("sample.xlsx")
    wb.save(output)

if __name__ == "__main__":

    table_to_excel(arcpy.GetParameter(0), arcpy.GetParameterAsText(1), arcpy.GetParameterAsText(2), arcpy.GetParameterAsText(3))
